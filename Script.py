import os
from googleapiclient.discovery import build
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
import time

# Google Sheets API credentials
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
CREDS_FILE = 'credentials.json'
SPREADSHEET_ID = 'your_spreadsheet_id'
RANGE_NAME = 'Sheet1!A1'

def get_credentials():
    creds = ServiceAccountCredentials.from_json_keyfile_name(CREDS_FILE, SCOPES)
    return creds

def update_google_sheet(values):
    creds = get_credentials()
    service = build('sheets', 'v4', credentials=creds)

    body = {
        'values': values
    }

    result = service.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID, range=RANGE_NAME,
        valueInputOption='RAW', body=body).execute()
    print('{0} cells updated.'.format(result.get('updatedCells')))

def read_excel_file(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

if __name__ == "__main__":
    excel_file_path = 'path_to_your_excel_file.xlsx'

    while True:
        if os.path.exists(excel_file_path):
            data = read_excel_file(excel_file_path)
            update_google_sheet(data)
        else:
            print("Excel file not found.")
        time.sleep(60)  # Check for updates every minute
